"""Extract text from all PDFs and XLSX files in a folder into <folder>/_text/.

Usage:
  python extract_reports.py <FOLDER>

Uses pdfplumber for PDFs (tables) and openpyxl for spreadsheets.
Part of the Romans report-processing pipeline (see workflows/romans-report-processing.md).
"""
import sys
from pathlib import Path

def pdf_text(path):
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        return "\n".join((pg.extract_text() or "") for pg in pdf.pages)

def xlsx_text(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"=== SHEET: {ws.title} ({ws.max_row}x{ws.max_column}) ===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 400:
                out.append("...(truncated)")
                break
            cells = [("" if c is None else str(c)) for c in row]
            if any(c.strip() for c in cells):
                out.append("\t".join(cells))
    return "\n".join(out)

def main():
    if len(sys.argv) < 2:
        print("usage: python extract_reports.py <FOLDER>")
        sys.exit(1)
    folder = Path(sys.argv[1])
    txt = folder / "_text"
    txt.mkdir(exist_ok=True)
    for f in sorted(folder.glob("*")):
        if f.suffix.lower() == ".pdf":
            t = pdf_text(f)
        elif f.suffix.lower() == ".xlsx":
            t = xlsx_text(f)
        else:
            continue
        (txt / (f.stem + ".txt")).write_text(t, encoding="utf-8")
        print(f"{len(t):>8} chars  {f.name}")
    print("DONE ->", txt)

if __name__ == "__main__":
    main()
